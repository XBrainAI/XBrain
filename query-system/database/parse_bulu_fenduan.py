import openpyxl

wb = openpyxl.load_workbook(r'd:\data\wy25311753\workspace\ws_workbuddy\ws_study8\database\2025年广州补录考生分段统计.xlsx')
ws = wb.active

# 读取表头（第3行）
headers = [ws.cell(3, c).value for c in range(1, ws.max_column + 1)]
print('Headers:', headers)

# 读取数据行（从第4行开始）
rows = []
for r in range(4, ws.max_row + 1):
    row_data = [ws.cell(r, c).value for c in range(1, ws.max_column + 1)]
    if any(v is not None for v in row_data):
        rows.append(row_data)

print(f'共 {len(rows)} 行数据')

# 生成HTML - 单表格结构
html_content = '''<!DOCTYPE html>
<html lang="zh-CN">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<title>2025年广州补录考生分数段统计</title>
<style>
  * { box-sizing: border-box; margin: 0; padding: 0; }
  body {
    font-family: "Microsoft YaHei", "PingFang SC", sans-serif;
    background: #f5f7fa;
    color: #1a1a2e;
    padding: 20px;
    line-height: 1.6;
  }
  .container {
    max-width: 1200px;
    margin: 0 auto;
    background: #fff;
    border-radius: 12px;
    box-shadow: 0 4px 20px rgba(0,0,0,0.08);
    padding: 32px;
  }
  h1 {
    text-align: center;
    font-size: 24px;
    margin-bottom: 8px;
    color: #1a1a2e;
  }
  .subtitle {
    text-align: center;
    color: #666;
    font-size: 14px;
    margin-bottom: 32px;
  }
  table {
    width: 100%;
    border-collapse: collapse;
    font-size: 14px;
  }
  th, td {
    border: 1px solid #e2e8f0;
    padding: 10px 12px;
    text-align: center;
  }
  th {
    background: #f1f5f9;
    font-weight: 600;
    color: #334155;
  }
  tr:nth-child(even) { background: #f8fafc; }
  tr:hover { background: #e0f2fe; }
  .highlight { font-weight: 700; color: #dc2626; }
  .num { font-family: "SF Mono", monospace; }
  .footer {
    text-align: center;
    color: #94a3b8;
    font-size: 12px;
    margin-top: 24px;
    padding-top: 16px;
    border-top: 1px solid #e2e8f0;
  }
</style>
</head>
<body>
<div class="container">
  <h1>2025年广州补录考生分数段统计</h1>
  <p class="subtitle">数据来源：2025年广州补录考生分段统计.xlsx</p>

  <table>
    <thead>
      <tr>
        <th>分数段</th>
        <th>全市累计考生数</th>
        <th>分数段</th>
        <th>全市段内考生数</th>
        <th>分数段</th>
        <th>全市段内考生数</th>
      </tr>
    </thead>
    <tbody>
'''

for row in rows:
    html_content += '      <tr>\n'
    for i, val in enumerate(row):
        if i in [1, 3, 5]:
            html_content += f'        <td class="num highlight">{val if val is not None else ""}</td>\n'
        else:
            html_content += f'        <td>{val if val is not None else ""}</td>\n'
    html_content += '      </tr>\n'

html_content += '''    </tbody>
  </table>

  <div class="footer">
    <p>数据整理时间：2025年 | 仅供志愿填报参考</p>
  </div>
</div>
</body>
</html>
'''

output_path = r'd:\data\wy25311753\workspace\ws_workbuddy\ws_study8\query-system\other_infos\2025年广州补录考生分数段统计.html'
with open(output_path, 'w', encoding='utf-8') as f:
    f.write(html_content)

print(f'已保存到: {output_path}')
