import openpyxl
from pathlib import Path

wb = openpyxl.load_workbook(r'd:\data\wy25311753\workspace\ws_workbuddy\ws_study8\database\2025年补录.xlsx')
ws = wb.active

# 找到表头行（第3行）
header_row = 3
headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
print('Headers:', headers)

records = []
for r in range(header_row + 1, ws.max_row + 1):
    idx = ws.cell(r, 1).value
    if idx is None:
        continue
    name = ws.cell(r, 2).value
    nature = ws.cell(r, 3).value
    plan = ws.cell(r, 4).value
    score = ws.cell(r, 5).value
    if nature not in ('公办', '民办'):
        continue
    records.append({
        '序号': idx,
        '学校名称': name,
        '学校性质': nature,
        '补录计划': plan,
        '补录最低控制分数线': score,
    })

print(f'共提取 {len(records)} 条记录')
for rec in records[:5]:
    print(rec)

# 生成 markdown
lines = [
    '# 2025年广州市普通高中和中本贯通补录计划与最低控制分数线',
    '',
    '> 数据来源：2025年补录.xlsx',
    '> 筛选条件：学校性质为「公办」或「民办」',
    '',
    '| 序号 | 学校名称 | 学校性质 | 补录计划 | 补录最低控制分数线 |',
    '|------|----------|----------|----------|--------------------|',
]
for rec in records:
    lines.append(f"| {rec['序号']} | {rec['学校名称']} | {rec['学校性质']} | {rec['补录计划']} | {rec['补录最低控制分数线']} |")

md_path = r'd:\data\wy25311753\workspace\ws_workbuddy\ws_study8\database\2025年补录计划与最低控制分数线.md'
with open(md_path, 'w', encoding='utf-8') as f:
    f.write('\n'.join(lines))

print(f'\n已保存到: {md_path}')
